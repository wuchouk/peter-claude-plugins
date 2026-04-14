#!/usr/bin/env python3
"""fill-excel.py — Excel 表單填寫：解密資料 + openpyxl 填寫

用法: fill-excel.py <xlsx_file> <mapping.json>

⚠️ stdout 只輸出狀態訊息，不輸出任何資料值
"""

import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook


def resolve_key(data, key):
    """解析 dot notation key，支援陣列 [index]"""
    parts = re.split(r'\.|\[(\d+)\]', key)
    parts = [p for p in parts if p is not None and p != '']
    current = data
    for part in parts:
        if part.isdigit():
            idx = int(part)
            if isinstance(current, list) and idx < len(current):
                current = current[idx]
            else:
                return None
        elif isinstance(current, dict) and part in current:
            current = current[part]
        else:
            return None
    return current


def apply_transform(value, transform):
    """套用值轉換"""
    if not transform or value is None:
        return value

    value = str(value)

    if transform.startswith("date_format:"):
        fmt = transform.split(":", 1)[1]
        try:
            dt = datetime.strptime(value, "%Y-%m-%d")
            result = fmt
            result = result.replace("YYYY", str(dt.year))
            result = result.replace("YY", str(dt.year)[-2:])
            result = result.replace("MM", f"{dt.month:02d}")
            result = result.replace("DD", f"{dt.day:02d}")
            result = result.replace("month_name", dt.strftime("%B"))
            result = result.replace("month_abbr", dt.strftime("%b"))
            return result
        except ValueError:
            return value

    if transform.startswith("date_part:"):
        part = transform.split(":", 1)[1]
        try:
            dt = datetime.strptime(value, "%Y-%m-%d")
            parts_map = {
                "year": str(dt.year),
                "month": f"{dt.month:02d}",
                "day": f"{dt.day:02d}",
                "month_name": dt.strftime("%B"),
                "month_abbr": dt.strftime("%b"),
            }
            return parts_map.get(part, value)
        except ValueError:
            return value

    if transform.startswith("select_map:"):
        map_str = transform.split(":", 1)[1].strip("{}")
        mapping = {}
        for pair in map_str.split(","):
            k, v = pair.split(":", 1)
            mapping[k.strip()] = v.strip()
        return mapping.get(value, mapping.get("*", value))

    if transform == "uppercase":
        return value.upper()
    if transform == "lowercase":
        return value.lower()
    if transform == "phone_format":
        return re.sub(r'[\s\-\(\)]', '', value)
    if transform == "phone_international":
        match = re.match(r'(\+\d{1,3})(0+)(.*)', value)
        if match:
            return match.group(1) + match.group(3)
        return value
    if transform.startswith("truncate:"):
        n = int(transform.split(":", 1)[1])
        return value[:n]

    return value


def decrypt_data(script_dir):
    """呼叫 encrypt.sh 解密，回傳暫存檔路徑（金鑰從 macOS Keychain 取得）"""
    cmd = ["bash", os.path.join(script_dir, "encrypt.sh"), "--decrypt"]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"❌ 解密失敗: {result.stderr.strip()}", file=sys.stderr)
        sys.exit(1)

    tmpfile = result.stdout.strip()
    if not os.path.exists(tmpfile):
        print("❌ 解密暫存檔不存在", file=sys.stderr)
        sys.exit(1)

    return tmpfile


def fill_excel(xlsx_path, mapping_path):
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 解密
    print("🔐 解密個人資料...")
    tmpfile = decrypt_data(script_dir)

    try:
        with open(tmpfile, "r") as f:
            data = yaml.safe_load(f)

        with open(mapping_path, "r") as f:
            mappings = json.load(f)

        # 讀取 Excel
        print(f"📊 讀取 Excel: {xlsx_path}")
        wb = load_workbook(xlsx_path)

        filled = 0
        errors = []

        for m in mappings:
            sheet_name = m["sheet"]
            cell_ref = m["cell"]
            key = m.get("key", "")
            transform = m.get("transform", "")

            # 找到 sheet
            if sheet_name not in wb.sheetnames:
                errors.append(f"Sheet 不存在: {sheet_name}")
                continue

            ws = wb[sheet_name]

            # 取得值
            if "keys" in m:
                values = [resolve_key(data, k) for k in m["keys"]]
                values = [str(v) for v in values if v is not None]
                if not values:
                    errors.append(f"資料 keys 不存在: {m['keys']}")
                    continue
                value = m.get("join", " ").join(values)
            else:
                value = resolve_key(data, key)
                if value is None:
                    errors.append(f"資料 key 不存在: {key}")
                    continue

            value = apply_transform(value, transform)
            if value is None:
                continue

            # 填入儲存格
            try:
                ws[cell_ref] = str(value)
                filled += 1
            except Exception as e:
                errors.append(f"{sheet_name}!{cell_ref}: {str(e)}")

        # 輸出
        output_path = xlsx_path.replace(".xlsx", "_filled.xlsx")
        wb.save(output_path)

        print(f"✅ 填入 {filled} 個欄位")
        if errors:
            print(f"⚠️  {len(errors)} 個錯誤:")
            for e in errors:
                print(f"   - {e}")
        print(f"📁 輸出: {output_path}")

    finally:
        # 安全刪除暫存檔
        if os.path.exists(tmpfile):
            subprocess.run(["rm", "-P", tmpfile], capture_output=True)
            if os.path.exists(tmpfile):
                os.remove(tmpfile)
            print("🗑️  暫存檔已清除")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: fill-excel.py <xlsx_file> <mapping.json>", file=sys.stderr)
        sys.exit(1)

    xlsx_file = sys.argv[1]
    mapping_file = sys.argv[2]
    if not os.path.exists(xlsx_file):
        print(f"❌ 找不到 Excel: {xlsx_file}", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(mapping_file):
        print(f"❌ 找不到映射檔: {mapping_file}", file=sys.stderr)
        sys.exit(1)

    fill_excel(xlsx_file, mapping_file)
